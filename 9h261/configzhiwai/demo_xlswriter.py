import datetime,os,sys,xlrd

import xlsxwriter
from openpyxl import load_workbook
from jira import JIRA

jira_server = 'http://jira.sunaw.com'
jira_username = 'zuowei'
jira_password = 'zw1997515'
# jira_username = 'xiaowenb'
# jira_password = 'xwb2834'

# jira = JIRA(basic_auth=(jira_username, jira_password), options = {'server': jira_server})
jira = JIRA(jira_server, basic_auth=(jira_username, jira_password))  # 创建jira连接


def searchIssues(jql, max_results=10000):
    ''' Search issues
    @param jql: JQL, str
    @param max_results: max results, int, default 100
    @return issues: result, zlist
    '''
    try:
        issues = jira.search_issues(jql, maxResults=max_results)
        return issues
    except Exception as e:
        print(e)


jql = '''
    issuetype = Bug AND reporter in (currentUser()) order by created DESC
'''
issues = searchIssues(jql)
# print(len(issues))
# todo assignee：经办人created: 创建时间creator: 创建人labels: 标签priorit: 优先级progress:project: 所示项目reporter: 报告人
# todo status: 状态summary: 问题描述worklog: 活动日志updated: 更新时间watches: 关注者comments: 评论resolution: 解决方案
# todo subtasks: 子任务issuelinks: 连接问题lastViewed: 最近查看时间

zlist_assignee_all = []
for issue in issues:  # todo 总表
    # print(format(issue.fields.created)[0:10]
    zlist_assignee_all.append(format(issue.fields.assignee))  # 经办人
    # print('{0}: {1}:{2}'.format(issue.key, issue.fields.summary,issue.fields.assignee))
# print('zlist_assignee_all',zlist_assignee_all)
# print(                                                                        )
# print('----------926项目总bug数量：%s,明细如下：'%len(zlist_assignee_all))
myset = set(zlist_assignee_all)  # myset是另外一个列表，里面的内容是myzlist里面的无重复 项
# for item in myset:
#     print("All the %s has found %s" %(item,zlist_assignee_all.count(item)))


zlist_assignee_18 = []
for issue in issues:  # todo 总表
    if str(format(issue.fields.created)[0:4]) == '2018':
        zlist_assignee_18.append(format(issue.fields.assignee))
# print('----------个人项目18年总bug数量：%s,明细如下：' % len(zlist_assignee_18))
myset = set(zlist_assignee_18)  # myset是另外一个列表，里面的内容是myzlist里面的无重复 项
# for item in myset:
#     print("All the %s has found %s" %(item,zlist_assignee_18.count(item)))

zlist_assignee_number = []  # bug 总量表
zlist_assignee_19 = []  # 月份表
zlist_assignee_month = []  # 月份bug表[]
zlist_created = []  # # 创办时间表
zlist_left_over = []  # 遗留数
zlist_left_over_lv = []
zlist_xlsx = []


def month_926(month):
    for issue in issues:  # 2019年分表
        if str(format(issue.fields.created)[0:4]) == '2019' and int(
                format(issue.fields.created)[5:7]) == month:  # 获取2019年
            zlist_assignee_19.append(format(issue.fields.assignee))  # 经办人

            # print(len(zlist_assignee_19))
            zlist_assignee_month.append(zlist_assignee_19)  # 添加所有月份的bug数量

            if format(issue.fields.status) != "已关闭":
                # print(format(issue.fields.summary), format(issue.fields.status))
                zlist_left_over.append(format(issue.fields.assignee))  # 添加月份遗留数量
    if len(zlist_left_over):
        zlist_left_over_lv.append(zlist_left_over)
    else:
        zlist_left_over_lv.append([])
    # print('月份为',i,zlist_assignee_19,len(zlist_assignee_19))
    if zlist_assignee_number:  # 不为空时表示为2月以上 将之前的和2月以商的相加
        len(zlist_assignee_19) + zlist_assignee_number[len(zlist_assignee_number) - 1]
        zlist_assignee_number.append(
            len(zlist_assignee_19) + zlist_assignee_number[len(zlist_assignee_number) - 1])  # 添加进入年度表
    else:  # 为空时表示为1月 将18年的和1月的相加
        zlist_assignee_number.append(len(zlist_assignee_18) + len(zlist_assignee_19))  # 添加进入年度表

    # print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
    # print(zlist_left_over)
    # print(zlist_left_over_lv)
    # print('当前月份为%s ' % i)
    # print('当前总bug数量为%s' % zlist_assignee_number[i - 1])
    # print('新增数量为%s' % (len(zlist_assignee_month[i - 1])))
    # print('解决数量为%s' % (len(zlist_assignee_month[i - 1]) - len(zlist_left_over)))
    if len(zlist_left_over_lv[i - 1]):
        # print('遗留数为%s' % len(zlist_left_over_lv[i - 1]))
        zlist_xlsx.append(len(zlist_left_over_lv[i - 1]))  # 遗留数
    else:
        # print('遗留数为0')
        zlist_xlsx.append(0)  # 遗留数
    zlist_xlsx.append(zlist_assignee_number[i - 1])  # 当前总bug数
    if len(zlist_left_over_lv[i - 1]):
        # print('遗留率为%s' % str(len(zlist_left_over_lv[i - 1]) / len(zlist_assignee_month[i - 1]) * 100)[0:4])
        lv = str(len(zlist_left_over_lv[i - 1]) / len(zlist_assignee_month[i - 1]) * 100)[0:4]
        zlist_xlsx.append(float(lv))  # 遗留lv
    else:
        # print('遗留率为0')
        zlist_xlsx.append(0)  # 遗留lv
    # print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

    zlist_xlsx.append(len(zlist_assignee_month[i - 1]))  # 新增数量
    zlist_xlsx.append((len(zlist_assignee_month[i - 1]) - len(zlist_left_over)))  # 解决数量

    zlist_assignee_19.clear()  # 清空月表,进入下一个月
    zlist_left_over.clear()
    # myset = set(zlist_assignee_19)  # myset是另外一个列表，里面的内容是myzlist里面的无重复 项
    # for item in myset:
    #     print("2019 the %s has found %s" % (item, zlist_assignee_19.count(item)))


for i in range(1, 13):  # 遍历当前月份i
    # print("i = ",i)
    month_926(i)

#  ---------------------------------------------------------------------------------------------------------------------

path_1 = os.path.abspath(os.path.dirname(__file__))
print (os.path.abspath(os.path.dirname(__file__)))
print(path_1+"\chart_line.xlsx")
# 创建一个excel
workbook = xlsxwriter.Workbook(path_1+"\chart_line.xlsx")
# 创建一个sheet
worksheet = workbook.add_worksheet('zuowei')
# worksheet = workbook.add_worksheet("bug_analysis")

# 自定义样式，加粗
bold = workbook.add_format({'bold': 1})
merge_format2 = workbook.add_format({
    # 'bold':     True,
    'border': 6,
    'align': 'center',  # 水平居中
    'valign': 'vcenter',  # 垂直居中
    'fg_color': 'yellow',  # 颜色填充
    'text_wrap': 1,
})
merge_format = workbook.add_format({
    'bold': True,
    'border': 6,
    'align': 'center',  # 水平居中
    'valign': 'vcenter',  # 垂直居中
    'text_wrap': 1,

    # 'fg_color': '#D7E4BC',#颜色填充
})

worksheet.merge_range('A1:A100', '', merge_format)  # 设置A列为间隔
worksheet.set_column('A:A', 3)  # 设置  A  的单元格宽度为12
worksheet.set_column('B:B', 10)  # 设置  A  的单元格宽度为12
worksheet.set_column('C:C', 12)  # 设置  A  的单元格宽度为12
worksheet.set_column('D:D', 10)  # 设置  A  的单元格宽度为12
worksheet.set_column('E:E', 10)  # 设置  A  的单元格宽度为12
worksheet.set_column('F:F', 15)  # 设置  A  的单元格宽度为12
worksheet.set_column('G:G', 10)  # 设置  A  的单元格宽度为12
worksheet.set_column('H:H', 10)  # 设置  A  的单元格宽度为12

worksheet.merge_range('B1:H1', '2019年个人年度提交缺陷状况表', merge_format)
worksheet.merge_range('I1:Z16', '', merge_format)
worksheet.merge_range('B17:Z100', '', merge_format)
worksheet.merge_range('B2:B3', '月份', merge_format)  # 月份
worksheet.merge_range('C2:C3', '统计时间', merge_format)  # 统计时间
worksheet.merge_range('D2:D3', '遗留数', merge_format2)
worksheet.merge_range('E2:E3', '总数', merge_format2)
worksheet.merge_range('F2:F3', '总数遗留率(%)', merge_format2)
worksheet.merge_range('G2:G3', '新增数', merge_format2)
worksheet.merge_range('H2:H3', '解决数', merge_format2)

# --------1、准备数据并写入excel---------------
# 向excel中写入数据，建立图标时要用到
zlist1 = []  # 遗留数
zlist2 = []
zlist3 = []
zlist4 = []
zlist5 = []

# 写入数据
for i in range(0, int(len(zlist_xlsx) / 5)):
    # print(a)
    # print("++++++++++")
    zlist1.append(zlist_xlsx[i * 5])
    zlist2.append(zlist_xlsx[i * 5 + 1])
    zlist3.append(zlist_xlsx[i * 5 + 2])
    zlist4.append(zlist_xlsx[i * 5 + 3])
    zlist5.append(zlist_xlsx[i * 5 + 4])

sum = 0
for i in zlist1:
    sum += i
zlist1.append(sum)
zlist2.append(zlist2[-1])

lv_all = sum / zlist2[-1]
zlist3.append(round(float(lv_all * 100), 2))

# print("+++++++++++++++++++++++++++++++++++++")
# print(zlist_xlsx)
# print("+++++++++++++++++++++++++++++++++++++")
data = [
    ['2019-01月', '2019-02月', '2019-03月', '2019-04月', '2019-05月', '2019-06月', '2019-07月', '2019-08月'
        , '2019-09月', '2019-10月', '2019-11月', '2019-12月', "年度汇总"],
    ['1月31日', '2月28日', '3月28日', '4月30日', '5月30日', '6月27日', '7月31日', '8月29日'
        , '9月30日', '10月31日', '11月28日', '12月26日',
     str(datetime.datetime.now().month) + "月" + str(datetime.datetime.now().day)+"日(当日)"], zlist1, zlist2, zlist3, zlist4,
    zlist5

]

# 写入表头
# worksheet.write_row('B1', bold)
# 写入数据
worksheet.write_column('B4', data[0])
worksheet.write_column('C4', data[1])
worksheet.write_column('D4', zlist1, merge_format2)
worksheet.write_column('E4', zlist2,merge_format2)
worksheet.write_column('F4', zlist3, merge_format2)
worksheet.write_column('G4', zlist4, merge_format2)
worksheet.write_column('H4', zlist5, merge_format2)

# worksheet.set_row(7, 30)#设置第8行的高度为30
# --------2、生成图表并插入到excel---------------
# 创建一个柱状图(line chart)
zchart_col = workbook.add_chart({'type': 'line'})

# 配置第一个系列数据
zchart_col.add_series({
    # 这里的zuowei是默认的值，因为我们在新建sheet时没有指定sheet名
    # 如果我们新建sheet时设置了sheet名，这里就要设置成相应的值
    'name': '=zuowei!$B$1',
    'categories': '=zuowei!$B$4:$B$15',  # X 轴
    'values': '=zuowei!$G$4:$G$15',  # Y
    'data_labels': {'value': True},
    'line': {'color': 'red'},
})

# # 配置第二个系列数据
# zchart_col.add_series({
#     'name': '=zuowei!$C$1',
#     'categories': '=zuowei!$A$2:$A$7',
#     'values': '=zuowei!$C$2:$C$7',
#     'line': {'color': 'yellow'},
# })

# 配置第二个系列数据(用了另一种语法)
# zchart_col.add_series({
#     'name': ['zuowei', 0, 2],
#     'categories': ['zuowei', 1, 0, 6, 0],
#     'values': ['zuowei', 1, 2, 6, 2],
#     'line': {'color': 'yellow'},
# })

# 设置图表的title 和 x，y轴信息
zchart_col.set_title({'name': '2019年个人年度提交缺陷状况表'})
zchart_col.set_x_axis({'name': '月份'})
zchart_col.set_y_axis({'name': '遗留率%'})

# 设置图表的风格
zchart_col.set_style(1)

# 把图表插入到worksheet并设置偏移
worksheet.insert_chart('B17', zchart_col, {'x_offset': 30, 'y_offset': 10})

#  todo--------------------------------------------------------------------------------------------------------------

jira_server = 'http://jira.sunaw.com'
jira_username = 'xiaowenb'
jira_password = 'xwb2834'

# jira = JIRA(basic_auth=(jira_username, jira_password), options = {'server': jira_server})
jira = JIRA(jira_server, basic_auth=(jira_username, jira_password))  # 创建jira连接


def searchIssues(jql, max_results=10000):
    ''' Search issues
    @param jql: JQL, str
    @param max_results: max results, int, default 100
    @return issues: result, zlist
    '''
    try:
        issues = jira.search_issues(jql, maxResults=max_results)
        return issues
    except Exception as e:
        print(e)


jql = '''
    issuetype = Bug AND reporter in (currentUser()) order by created DESC
'''
issues = searchIssues(jql)
# print(len(issues))
# todo assignee：经办人created: 创建时间creator: 创建人labels: 标签priorit: 优先级progress:project: 所示项目reporter: 报告人
# todo status: 状态summary: 问题描述worklog: 活动日志updated: 更新时间watches: 关注者comments: 评论resolution: 解决方案
# todo subtasks: 子任务issuelinks: 连接问题lastViewed: 最近查看时间

xlist_assignee_all = []
for issue in issues:  # todo 总表
    # print(format(issue.fields.created)[0:10]
    xlist_assignee_all.append(format(issue.fields.assignee))  # 经办人
    # print('{0}: {1}:{2}'.format(issue.key, issue.fields.summary,issue.fields.assignee))
# print('xlist_assignee_all',xlist_assignee_all)
# print(                                                                        )
# print('----------926项目总bug数量：%s,明细如下：'%len(xlist_assignee_all))
myset = set(xlist_assignee_all)  # myset是另外一个列表，里面的内容是myxlist里面的无重复 项
# for item in myset:
#     print("All the %s has found %s" %(item,xlist_assignee_all.count(item)))


xlist_assignee_18 = []
for issue in issues:  # todo 总表
    if str(format(issue.fields.created)[0:4]) == '2018':
        xlist_assignee_18.append(format(issue.fields.assignee))
# print('----------个人项目18年总bug数量：%s,明细如下：' % len(xlist_assignee_18))
myset = set(xlist_assignee_18)  # myset是另外一个列表，里面的内容是myxlist里面的无重复 项
# for item in myset:
#     print("All the %s has found %s" %(item,xlist_assignee_18.count(item)))

xlist_assignee_number = []  # bug 总量表
xlist_assignee_19 = []  # 月份表
xlist_assignee_month = []  # 月份bug表[]
xlist_created = []  # # 创办时间表
xlist_left_over = []  # 遗留数
xlist_left_over_lv = []
xlist_xlsx = []


def month_926(month):
    for issue in issues:  # 2019年分表
        if str(format(issue.fields.created)[0:4]) == '2019' and int(
                format(issue.fields.created)[5:7]) == month:  # 获取2019年
            xlist_assignee_19.append(format(issue.fields.assignee))  # 经办人

            # print(len(xlist_assignee_19))
            xlist_assignee_month.append(xlist_assignee_19)  # 添加所有月份的bug数量

            if format(issue.fields.status) != "已关闭":
                # print(format(issue.fields.summary), format(issue.fields.status))
                xlist_left_over.append(format(issue.fields.assignee))  # 添加月份遗留数量
    if len(xlist_left_over):
        xlist_left_over_lv.append(xlist_left_over)
    else:
        xlist_left_over_lv.append([])
    # print('月份为',i,xlist_assignee_19,len(xlist_assignee_19))
    if xlist_assignee_number:  # 不为空时表示为2月以上 将之前的和2月以商的相加
        len(xlist_assignee_19) + xlist_assignee_number[len(xlist_assignee_number) - 1]
        xlist_assignee_number.append(
            len(xlist_assignee_19) + xlist_assignee_number[len(xlist_assignee_number) - 1])  # 添加进入年度表
    else:  # 为空时表示为1月 将18年的和1月的相加
        xlist_assignee_number.append(len(xlist_assignee_18) + len(xlist_assignee_19))  # 添加进入年度表
    if len(xlist_left_over_lv[i - 1]):
        # print('遗留数为%s' % len(xlist_left_over_lv[i - 1]))
        xlist_xlsx.append(len(xlist_left_over_lv[i - 1]))  # 遗留数
    else:
        # print('遗留数为0')
        xlist_xlsx.append(0)  # 遗留数
    xlist_xlsx.append(xlist_assignee_number[i - 1])  # 当前总bug数
    if len(xlist_left_over_lv[i - 1]):
        # print('遗留率为%s' % str(len(xlist_left_over_lv[i - 1]) / len(xlist_assignee_month[i - 1]) * 100)[0:4])
        lv = str(len(xlist_left_over_lv[i - 1]) / len(xlist_assignee_month[i - 1]) * 100)[0:4]
        xlist_xlsx.append(float(lv))  # 遗留lv
    else:
        # print('遗留率为0')
        xlist_xlsx.append(0)  # 遗留lv
    # print("++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++++")

    xlist_xlsx.append(len(xlist_assignee_month[i - 1]))  # 新增数量
    xlist_xlsx.append((len(xlist_assignee_month[i - 1]) - len(xlist_left_over)))  # 解决数量

    xlist_assignee_19.clear()  # 清空月表,进入下一个月
    xlist_left_over.clear()


for i in range(1, 13):  # 遍历当前月份i
    # print("i = ",i)
    month_926(i)

# 创建一个sheet
worksheet = workbook.add_worksheet('xiaowenbo')
# worksheet = workbook.add_worksheet("bug_analysis")

# 自定义样式，加粗
bold = workbook.add_format({'bold': 1})
merge_format2 = workbook.add_format({
    # 'bold':     True,
    'border': 6,
    'align': 'center',  # 水平居中
    'valign': 'vcenter',  # 垂直居中
    'fg_color': 'yellow',  # 颜色填充
    'text_wrap': 1,
})
merge_format = workbook.add_format({
    'bold': True,
    'border': 6,
    'align': 'center',  # 水平居中
    'valign': 'vcenter',  # 垂直居中
    'text_wrap': 1,

    # 'fg_color': '#D7E4BC',#颜色填充
})
worksheet.merge_range('A1:A100', '', merge_format)  # 设置A列为间隔
worksheet.set_column('A:A', 3)  # 设置  A  的单元格宽度为12
worksheet.set_column('B:B', 10)  # 设置  A  的单元格宽度为12
worksheet.set_column('C:C', 12)  # 设置  A  的单元格宽度为12
worksheet.set_column('D:D', 10)  # 设置  A  的单元格宽度为12
worksheet.set_column('E:E', 10)  # 设置  A  的单元格宽度为12
worksheet.set_column('F:F', 15)  # 设置  A  的单元格宽度为12
worksheet.set_column('G:G', 10)  # 设置  A  的单元格宽度为12
worksheet.set_column('H:H', 10)  # 设置  A  的单元格宽度为12

worksheet.merge_range('B1:H1', '2019年个人年度提交缺陷状况表', merge_format)
worksheet.merge_range('I1:Z16', '', merge_format)
worksheet.merge_range('B17:Z100', '', merge_format)
worksheet.merge_range('B2:B3', '月份', merge_format)  # 月份
worksheet.merge_range('C2:C3', '统计时间', merge_format)  # 统计时间
worksheet.merge_range('D2:D3', '遗留数', merge_format2)
worksheet.merge_range('E2:E3', '总数', merge_format2)
worksheet.merge_range('F2:F3', '总数遗留率(%)', merge_format2)
worksheet.merge_range('G2:G3', '新增数', merge_format2)
worksheet.merge_range('H2:H3', '解决数', merge_format2)

# --------1、准备数据并写入excel---------------
# 向excel中写入数据，建立图标时要用到
xlist1 = []  # 遗留数
xlist2 = []
xlist3 = []
xlist4 = []
xlist5 = []

# 写入数据
for i in range(0, int(len(xlist_xlsx) / 5)):
    # print(a)
    # print("++++++++++")
    xlist1.append(xlist_xlsx[i * 5])
    xlist2.append(xlist_xlsx[i * 5 + 1])
    xlist3.append(xlist_xlsx[i * 5 + 2])
    xlist4.append(xlist_xlsx[i * 5 + 3])
    xlist5.append(xlist_xlsx[i * 5 + 4])

sum = 0
for i in xlist1:
    sum += i
xlist1.append(sum)
xlist2.append(xlist2[-1])

lv_all = sum / xlist2[-1]
xlist3.append(round(float(lv_all * 100), 2))

# print("+++++++++++++++++++++++++++++++++++++")
# print(xlist_xlsx)
# print("+++++++++++++++++++++++++++++++++++++")
data = [
    ['2019-01月', '2019-02月', '2019-03月', '2019-04月', '2019-05月', '2019-06月', '2019-07月', '2019-08月'
        , '2019-09月', '2019-10月', '2019-11月', '2019-12月', "年度汇总"],
    ['1月31日', '2月28日', '3月28日', '4月30日', '5月30日', '6月27日', '7月31日', '8月29日'
        , '9月30日', '10月31日', '11月28日', '12月26日',
     str(datetime.datetime.now().month) + "月" + str(datetime.datetime.now().day)+"日(当日)"],xlist1,xlist2,xlist3,xlist4,xlist5

]

# 写入表头
# worksheet.write_row('B1', bold)
# 写入数据
worksheet.write_column('B4', data[0])
worksheet.write_column('C4', data[1])
worksheet.write_column('D4', xlist1, merge_format2)
worksheet.write_column('E4', xlist2, merge_format2)
worksheet.write_column('F4', xlist3, merge_format2)
worksheet.write_column('G4', xlist4, merge_format2)
worksheet.write_column('H4', xlist5, merge_format2)

# worksheet.set_row(7, 30)#设置第8行的高度为30
# --------2、生成图表并插入到excel---------------
# 创建一个柱状图(line chart)
xchart_col = workbook.add_chart({'type': 'line'})

# 配置第一个系列数据
xchart_col.add_series({
    # 这里的zuowei是默认的值，因为我们在新建sheet时没有指定sheet名
    # 如果我们新建sheet时设置了sheet名，这里就要设置成相应的值
    'name': '=xiaowenbo!$B$1',
    'categories': '=xiaowenbo!$B$4:$B$15',  # X 轴
    'values': '=xiaowenbo!$G$4:$G$15',  # Y
    'data_labels': {'value': True},
    'line': {'color': 'red'},
})

# # 配置第二个系列数据
# xchart_col.add_series({
#     'name': '=xiaowenbo!$C$1',
#     'categories': '=xiaowenbo!$B2:$B$15',
#     'values': '=xiaowenbo!$B$2:$C$7',
#     'line': {'color': 'yellow'},
# })

# 配置第二个系列数据(用了另一种语法)
# xchart_col.add_series({
#     'name': ['xiaowenbo', 0, 2],
#     'categories': ['xiaowenbo', 1, 0, 6, 0],
#     'values': ['xiaowenbo', 1, 2, 6, 2],
#     'line': {'color': 'yellow'},
# })

# 设置图表的title 和 x，y轴信息
xchart_col.set_title({'name': '2019年个人年度提交缺陷状况表'})
xchart_col.set_x_axis({'name': '月份'})
xchart_col.set_y_axis({'name': '遗留率%'})

# 设置图表的风格
xchart_col.set_style(1)

# 把图表插入到worksheet并设置偏移
worksheet.insert_chart('B17', xchart_col, {'x_offset': 30, 'y_offset': 12})


# wb = load_workbook(path_1+"\chart_line.xlsx")#文件路径


workbook.close()
